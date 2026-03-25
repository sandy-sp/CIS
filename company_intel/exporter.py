from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from company_intel.models import CrawlJob, ExtractedEntity, PageRecord


_HEADER_BG = "203764"
_HEADER_FG = "FFFFFF"
_ALT_BG = "F3F6FB"
_LINK = "2E75B6"


def _set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _header(cell, value: str) -> None:
    cell.value = value
    cell.font = Font(bold=True, color=_HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _alt_fill(ws, row: int, cols: int) -> None:
    if row % 2 != 0:
        return
    fill = PatternFill("solid", fgColor=_ALT_BG)
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).fill = fill


def _link_cell(ws, row: int, col: int, text: str, url: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    if url:
        cell.hyperlink = url
        cell.font = Font(color=_LINK, underline="single")


class IntelExporter:
    def write_excel(self, job: CrawlJob, records: list[PageRecord],
                    entities: dict[str, list[ExtractedEntity]], output_path: Path) -> Path:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._write_summary(wb, job, records, entities)
        self._write_simple_sheet(wb, "Services", entities.get("services", []),
                                 ["Name", "Summary", "Source URL", "Confidence"],
                                 lambda entity: [entity.display_name, entity.attributes.get("summary", ""),
                                                 entity.source_urls[0] if entity.source_urls else "", entity.confidence],
                                 [32, 70, 55, 14])
        self._write_simple_sheet(wb, "Case Studies", entities.get("case_studies", []),
                                 ["Name", "Summary", "Source URL", "Confidence"],
                                 lambda entity: [entity.display_name, entity.attributes.get("summary", ""),
                                                 entity.source_urls[0] if entity.source_urls else "", entity.confidence],
                                 [36, 70, 55, 14])
        self._write_simple_sheet(wb, "Partners", entities.get("partners", []),
                                 ["Name", "Category", "Source URL", "Confidence"],
                                 lambda entity: [entity.display_name, entity.attributes.get("category", ""),
                                                 entity.source_urls[0] if entity.source_urls else "", entity.confidence],
                                 [32, 28, 55, 14])
        self._write_simple_sheet(wb, "Customers", entities.get("customers", []),
                                 ["Name", "Context", "Source URL", "Confidence"],
                                 lambda entity: [entity.display_name, entity.attributes.get("context", ""),
                                                 entity.source_urls[0] if entity.source_urls else "", entity.confidence],
                                 [32, 70, 55, 14])
        self._write_simple_sheet(wb, "People", entities.get("people", []),
                                 ["Name", "Title", "LinkedIn URL", "Source URL", "Confidence"],
                                 lambda entity: [
                                     entity.display_name,
                                     entity.attributes.get("title", ""),
                                     entity.attributes.get("linkedin_url", ""),
                                     entity.source_urls[0] if entity.source_urls else "",
                                     entity.confidence,
                                 ],
                                 [28, 28, 55, 55, 14])
        self._write_simple_sheet(wb, "Events", entities.get("events", []),
                                 ["Name", "Date", "Location", "Event Type", "Source URL", "Confidence"],
                                 lambda entity: [
                                     entity.display_name,
                                     entity.attributes.get("date", ""),
                                     entity.attributes.get("location", ""),
                                     entity.attributes.get("event_type", ""),
                                     entity.source_urls[0] if entity.source_urls else "",
                                     entity.confidence,
                                 ],
                                 [38, 20, 20, 18, 55, 14])
        self._write_inventory(wb, records)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    def _write_summary(self, wb, job: CrawlJob, records: list[PageRecord],
                       entities: dict[str, list[ExtractedEntity]]) -> None:
        ws = wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False
        _set_widths(ws, [32, 18])
        rows = [
            ("Domain", job.domain),
            ("Status", job.status),
            ("Pages scraped", job.pages_scraped),
            ("Pages failed", job.pages_failed),
            ("Pages skipped", job.pages_skipped),
            ("Access denied", job.pages_blocked),
            ("Total words", sum(record.word_count for record in records if record.status == "success")),
            ("Services", len(entities.get("services", []))),
            ("Case Studies", len(entities.get("case_studies", []))),
            ("Partners", len(entities.get("partners", []))),
            ("Customers", len(entities.get("customers", []))),
            ("People", len(entities.get("people", []))),
            ("Events", len(entities.get("events", []))),
        ]
        _header(ws.cell(row=1, column=1), "Metric")
        _header(ws.cell(row=1, column=2), "Value")
        for idx, (label, value) in enumerate(rows, start=2):
            ws.cell(row=idx, column=1, value=label)
            ws.cell(row=idx, column=2, value=value)
            _alt_fill(ws, idx, 2)
        ws.freeze_panes = "A2"

    def _write_simple_sheet(self, wb, name: str, entities: list[ExtractedEntity], headers, row_fn, widths) -> None:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        _set_widths(ws, widths)
        for idx, header_name in enumerate(headers, start=1):
            _header(ws.cell(row=1, column=idx), header_name)
        for row, entity in enumerate(entities, start=2):
            values = row_fn(entity)
            for col, value in enumerate(values, start=1):
                if headers[col - 1].endswith("URL"):
                    _link_cell(ws, row, col, value, value)
                else:
                    ws.cell(row=row, column=col, value=value)
            _alt_fill(ws, row, len(headers))
        ws.freeze_panes = "A2"

    def _write_inventory(self, wb, records: list[PageRecord]) -> None:
        ws = wb.create_sheet("Page Inventory")
        ws.sheet_view.showGridLines = False
        headers = ["URL", "Category", "Subtype", "Source Type", "Status", "Words", "Engine", "Duplicate", "Source File"]
        widths = [64, 18, 18, 14, 12, 12, 14, 12, 40]
        _set_widths(ws, widths)
        for idx, header_name in enumerate(headers, start=1):
            _header(ws.cell(row=1, column=idx), header_name)
        for row, record in enumerate(sorted(records, key=lambda item: item.url), start=2):
            _link_cell(ws, row, 1, record.url, record.url)
            ws.cell(row=row, column=2, value=record.page_category)
            ws.cell(row=row, column=3, value=record.page_subtype)
            ws.cell(row=row, column=4, value=record.source_type)
            ws.cell(row=row, column=5, value=record.status)
            ws.cell(row=row, column=6, value=record.word_count)
            ws.cell(row=row, column=7, value=record.engine_used)
            ws.cell(row=row, column=8, value="yes" if record.is_duplicate else "")
            ws.cell(row=row, column=9, value=record.source_file)
            _alt_fill(ws, row, len(headers))
        ws.freeze_panes = "A2"
